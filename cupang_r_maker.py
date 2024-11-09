# streamlit run cupang_r_maker.py
import streamlit as st
#multiprocess error
from multiprocessing import freeze_support

from langchain_google_genai import ChatGoogleGenerativeAI #LLM Setting
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser 

from bs4 import BeautifulSoup as bs
from typing import Optional,Union,Dict,List
from openpyxl import Workbook
import time
import os
import re
import requests as rq
import pandas as pd

#Steamlit cloud에서 GitHub 아이콘 숨기기용 CSS 코드
hide_code = """
<style>
.streamlit-expanderHeader {
    display: none;
}
</style>
"""
st.markdown(hide_code, unsafe_allow_html=True)
#######################################################

def get_headers(
    key: str,
    default_value: Optional[str] = None
    )-> Dict[str,Dict[str,str]]:
    """ Get Headers """
    jsonx = {"headers": {
        "authority": "weblog.coupang.com",
        "scheme": "https",
        "origin": "https://www.coupang.com",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "macOS",
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Whale/3.20.182.14 Safari/537.36",
        "cookie": "PCID=31489593180081104183684; _fbp=fb.1.1644931520418.1544640325; gd1=Y; X-CP-PT-locale=ko_KR; MARKETID=31489593180081104183684; sid=03ae1c0ed61946c19e760cf1a3d9317d808aca8b; overrideAbTestGroup=%5B%5D; x-coupang-origin-region=KOREA; x-coupang-accept-language=ko_KR;"
        }
        }
    headers : Dict[str,Dict[str,str]] = jsonx

    try :
        return headers[key]
    except:
        if default_value:
            return default_value
        raise EnvironmentError(f'Set the {key}')

class Coupang:
    @staticmethod
    def get_product_code(url: str)-> str:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code : str = url.split('products/')[-1].split('?')[0]
        return prod_code

    def __init__(self)-> None:
        self.__headers : Dict[str,str] = get_headers(key='headers')

    def main(self)-> List[List[Dict[str,Union[str,int]]]]:
        # URL 주소
        URL : str = self.input_review_url()

        # URL의 Product Code 추출
        prod_code : str = self.get_product_code(url=URL)

        # URL 주소 재가공
        #URLS : List[str] = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true' for page in range(1,self.input_page_count() + 1)]
        URLS : List[str] = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true' for page in range(1,2)]

        # __headers에 referer 키 추가
        self.__headers['referer'] = URL

        with rq.Session() as session:
            return [self.fetch(url=url,session=session) for url in URLS]

    def fetch(self,url:str,session)-> List[Dict[str,Union[str,int]]]:
        save_data : List[Dict[str,Union[str,int]]] = list()

        with session.get(url=url,headers=self.__headers) as response :
            html = response.text
            soup = bs(html,'html.parser')

            # Article Boxes
            article_lenth = len(soup.select('article.sdp-review__article__list'))

            for idx in range(article_lenth):
                dict_data : Dict[str,Union[str,int]] = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 구매자 이름
                user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
                if user_name == None or user_name.text == '':
                    user_name = '-'
                else:
                    user_name = user_name.text.strip()

                # 평점
                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                if rating == None:
                    rating = 0
                else :
                    rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = '등록된 헤드라인이 없습니다'
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                if review_content == None :
                    review_content = '등록된 리뷰내용이 없습니다'
                else:
                    review_content = re.sub('[\n\t]','',review_content.text.strip())

                # 맛 만족도
                answer = articles[idx].select_one('span.sdp-review__article__list__survey__row__answer')
                if answer == None or answer.text == '':
                    answer = '맛 평가 없음'
                else:
                    answer = answer.text.strip()

                dict_data['prod_name'] = prod_name
                dict_data['user_name'] = user_name
                dict_data['rating'] = rating
                dict_data['headline'] = headline
                dict_data['review_content'] = review_content
                dict_data['answer'] = answer

                save_data.append(dict_data)

            time.sleep(1)

            return save_data
    
    @staticmethod
    def clear_console() -> None:
        command: str = 'clear'
        if os.name in ('nt','dos'):
            command = 'cls'
        os.system(command=command)

    def input_review_url(self)-> str:
        while True:
            self.clear_console()
            
            # Review URL
            #org review_url : str = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:')
            review_url : str = st.text_input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=:  ')
            #org review_url : str = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:')
            
            if not review_url :
                # Window
                os.system('cls')
                # Mac
                #os.system('clear')
                #org print('URL 주소가 입력되지 않았습니다')
                st.warning('URL 주소가 입력되지 않았습니다')
                continue
            return review_url

    def input_page_count(self)-> int:
        self.clear_console()

        while True:
            #org page_count : str = input('페이지 수를 입력하세요\n\n:')
            page_count : str = st.text_input('참고할 리뷰 페이지를 입력하세요:(숫자) ')
            if not page_count:
                #org print('페이지 수가 입력되지 않았습니다\n')
                st.warning('페이지 수가 입력되지 않았습니다.')
                continue

            return int(page_count)

class OpenPyXL:
    @staticmethod
    def save_file()-> None:
        # 크롤링 결과
        results : List[List[Dict[str,Union[str,int]]]] = Coupang().main()

        wb = Workbook()
        ws = wb.active
        ws.append(['상품명','구매자 이름','구매자 평점','리뷰 제목','리뷰 내용','맛 만족도'])

        row = 2

        for x in results:

            reviews = []

            for result in x :
                ws[f'A{row}'] = result['prod_name']
                ws[f'B{row}'] = result['user_name']
                ws[f'C{row}'] = result['rating']
                ws[f'D{row}'] = result['headline']
                ws[f'E{row}'] = result['review_content']
                ws[f'F{row}'] = result['answer']

                reviews.append(result['review_content'])

                row += 1
        
        df = pd.Series(data=reviews)

        return df

#Gemini API Key Setting
os.environ["GOOGLE_API_KEY"] = 'AIzaSyDRmcCNGKkn0ZfacIIaqQwGM-ZZZ27nmpw' ##enjin key_240927(new)

def review_maker(prod,ex,selected_model,num=500):

    # LCEL chaining
    chain = (
        ChatPromptTemplate.from_template(
        """
        당신은 쿠팡에서 {product}을 구매한 후에 {product}을 사용해 보고 {product}의 리뷰를 작성하는 인공지능 챗봇입니다.
        리뷰는 예제의 내용을 참고해서 만들고 또한 규칙을 지켜서 작성해야 합니다.

        [규칙]
        1) 리뷰는 반드시 아래 문장으로 시작해야 합니다.
        - 이번에 구매하게 된 {product}에 대한 사용기에 대해 말씀드릴께요.
        2) 전체 글자수는 {num}자 이하로 작성해야 합니다.
        3) 남자 아이 두명(중학교 1학년,초등학생)을 키우는 40대 초반의 주부라고 생각하고 작성해야 합니다.
        4) 장점과 단점을 구분해서 작성해야 합니다.
        
        [예제]
        1) {ex0}
        2) {ex1}
        3) {ex2}
        """
        ) 
        | ChatGoogleGenerativeAI(model=selected_model, temperature = 0.1) 
        | StrOutputParser()
    )
    # chain 호출
    resonse = chain.invoke({"product": prod, "ex0": ex[0], "ex1": ex[1],  "ex2": ex[2], "num": num})
    st.write(resonse)

def review_maker2(prod,selected_model,num=500):

    # LCEL chaining
    chain = (
        ChatPromptTemplate.from_template(
        """
        당신은 쿠팡에서 {product}을 구매한 후에 {product}을 사용해 보고 {product}의 리뷰를 작성하는 인공지능 챗봇입니다.
        리뷰는 규칙을 지켜서 작성해야 합니다.

        [규칙]
        1) 리뷰는 반드시 아래 문장으로 시작해야 합니다.
        - 이번에 구매하게 된 {product}에 대한 사용기에 대해 말씀드릴께요.
        2) 전체 글자수는 {num}자 이하로 작성해야 합니다.
        3) 남자 아이 두명(중학교 1학년,초등학생)을 키우는 40대 초반의 주부라고 생각하고 작성해야 합니다.
        4) 장점과 단점을 구분해서 작성해야 합니다.
        
        """
        ) 
        | ChatGoogleGenerativeAI(model=selected_model, temperature = 0.1) 
        | StrOutputParser()
    )
    # chain 호출
    resonse = chain.invoke({"product": prod, "num": num})
    st.write(resonse)
    

@st.cache_data()
def hold(hold_v):
    if(hold_v == 'OK'):
        st.write('사용자님 환영합니다.')
    else:
        st.error("잘못된 입력입니다. 다시 입력해 주세요.")
        st.stop()

if __name__ == "__main__":
    freeze_support() # for multiprocessing other process on windows
    
    
    # 페이지 기본 설정
    #st.set_page_config(
    #page_title="🔎쿠팡 리뷰 Maker",
    #layout="wide",
    #initial_sidebar_state="expanded")

    hold_v = 'No'

    with st.sidebar:

        random_val = int(st.text_input('문제를 선택하시오 0~10번 선택'))

        if random_val==0:
            namex = st.text_input('24년 9월 기준 김여사가 다니는 직장은? 9글자 ㅇㅇㅇㅇㅇㅇㅇ병원')
            if process :=st.button("Process"):
                if (namex != '세종충남대학교병원'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'
            
        if random_val==1:
            namex = st.text_input('24년 9월 기준 김여사가 가장 좋아하는 배우는? 3글자 김ㅇㅇ')
            if process :=st.button("Process"):
                if (namex != '김혜윤'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==2:
            namex = st.text_input('24년 9월 기준 김여사가 다니는 교회는? 6글자 ㅇㅇㅇㅇ교회')
            if process :=st.button("Process"):
                if (namex != '세종한빛교회'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==3:
            namex = st.text_input('24년 9월 기준 김여사의 사는 동이름은? 3글자 ㅇㅇ동')
            if process :=st.button("Process"):
                if (namex != '나성동'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==4:
            namex = st.text_input('24년 봄 김여사가 여행한 나라는? 2글자 ㅇㅇ')
            if process :=st.button("Process"):
                if (namex != '대만'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'
            
        if random_val==5:
            namex = st.text_input('김여사의 생년월일은? 숫자 8개만 입력 19ㅇㅇㅇㅇㅇㅇ')
            if process :=st.button("Process"):
                if (namex != '19810609'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==6:
            namex = st.text_input('24년 9월 기준 김여사가 가장 좋아하는 드라마는? 6글자 ㅇㅇㅇㅇㅇㅇ')
            if process :=st.button("Process"):
                if (namex != '선재업고튀어'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==7:
            namex = st.text_input('24년 9월 기준 김여사의 직업은? 7글자 ㅇㅇㅇㅇ연구원')
            if process :=st.button("Process"):
                if (namex != '임상배아연구원'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==8:
            namex = st.text_input('김여사가 고등학교 시절 좋아했던 가수는? 3글자 ㅇㅇㅇ')
            if process :=st.button("Process"):
                if (namex != '김종서'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'            

        if random_val==9:
            namex = st.text_input('김여사가 목장모임을 하는 장소는? 5글자 ㅇㅇㅇ센터')
            if process :=st.button("Process"):
                if (namex != '놀뛰날센터'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

        if random_val==10:
            namex = st.text_input('김여사의 고등학교 시절 최고의 발명품은? 4글자 ㅇㅇ집게')
            if process :=st.button("Process"):
                if (namex != '꼼꼼집게'):
                    st.error("잘못된 사용자입니다. 다시 입력해 주세요.")
                    hold_v = 'NO'
                    st.stop()
                else:
                    hold_v = 'OK'

    st.header("**:blue[_쿠팡_] :red[_Revies_] Maker** :sunglasses:")
    st.subheader(f"쿠팡 리뷰를 검색하고 AI로 리뷰를 작성합니다.")

    if namex:
        hold_v = 'OK'

    st.write(hold_v)

    hold(hold_v)

    #OpenPyXL.save_file()
    df = OpenPyXL.save_file()
       
    reviews = df.to_list()

    ex = [' ',' ',' ',' ', ' ']

    for index,i in enumerate(reviews):
       ex[index]=i

    selected_model = st.radio('Choose Gemini Model', ['gemini-1.5-flash', 'gemini-1.5-flash-latest','gemini-1.5-pro', 'gemini-1.5-pro-latest'], key='selected_model')

    st.info(f'{selected_model}을 선택하셨습니다.')
    
    selected_option = st.radio(
        '리뷰 종류를 선택하세요.:',('일반', '체험단',))

    if selected_option == '일반':
        num = 500
    elif selected_option == '체험단':
        num = 800

    if prod := st.text_input('제품명을 입력하세요. 일반단어로 표현해 주세요 ex) 샘표간장(x), 간장(o) >>>   '):

        
        st.info(f'{selected_option}을 선택하셔서 {num}자의 리뷰를 생성합니다.')

        st.subheader("기존 Review를 참고하여 작성한 리뷰입니다.")
        review_maker(prod,ex,selected_model,num)
        st.subheader("기존 Review를 참고하지 않고 작성한 리뷰입니다.")
        review_maker2(prod,selected_model,num) 
        st.info("리뷰 생성이 완료 됐습니다.")


    #1) 리뷰는 반드시 아래 2개의 문장으로 시작해야 합니다.
    #- 안녕하세요. 먹을 거 좋아하는 남자애 2명(초등 3학년, 중학 1학년)을 키우는 주부입니다.
    #- 이번에 구매하게 된 {product}에 대한 사용기에 대해 말씀드릴께요.
    #2) 전체 글자수는 {num}자 이하로 작성해야 합니다.
    #3) 남자 아이 두명(중학교 1학년,초등학생)을 키우는 40대 초반의 주부라고 생각하고 작성해야 합니다.
    #5) 문장과 문장사이는 구분이 되게 줄바꿈이 될 수 있도록 작성해야 합니다.
    #6) 장점과 단점을 구분해서 작성해야 합니다.

    #[규칙]
    #1) 리뷰는 반드시 아래 2개의 문장으로 시작해야 합니다.
    #- 안녕하세요. 먹을 거 좋아하는 남자애 2명(초등 3학년, 중학 1학년)을 키우는 주부입니다.
    #- 이번에 구매하게 된 {product}에 대한 사용기에 대해 말씀드릴께요.
    #2) 전체 글자수는 400자 이상 500자 이하로 작성해야 합니다.
    #3) 남자 아이 두명(중학교 1학년,초등학생)을 키우는 40대 초반의 주부라고 생각하고 작성해야 합니다.
    #5) 문장과 문장사이는 구분이 되게 줄바꿈이 될 수 있도록 작성해야 합니다.
    #6) 장점과 단점을 구분해서 작성해야 합니다.